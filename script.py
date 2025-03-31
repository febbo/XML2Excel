import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook

# Get the directory where the script is located
script_dir = os.path.dirname(os.path.abspath(__file__))

# Prompt user for the XML filename and build the full path
xml_file = input("Inserisci il nome del file XML (deve essere nella stessa cartella): ")
xml_path = os.path.join(script_dir, xml_file)

# Generate the Excel output filename and path based on the XML filename
excel_output = os.path.splitext(xml_file)[0] + ".xlsx"
excel_path = os.path.join(script_dir, excel_output)

# Parse the XML file
tree = ET.parse(xml_path)
root = tree.getroot()

# Create a new Excel workbook and remove the default sheet
wb = Workbook()
wb.remove(wb.active)

# Process each second-level element in the XML
for second_level in root:
    # Use the tag name as sheet name (limited to 31 characters for Excel compatibility)
    sheet_name = second_level.tag[:31]

    # Handle duplicate sheet names by adding a counter suffix
    counter = 1
    base_name = sheet_name
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_name}_{counter}"[:31]
        counter += 1

    # Create a new worksheet with the processed name
    ws = wb.create_sheet(title=sheet_name)

    # Collect all unique column names from the records
    column_names = set()
    for record in second_level.findall("record"):
        for column in record.findall("column"):
            if "name" in column.attrib:
                column_names.add(column.attrib["name"])

    # Sort column names for consistent output
    column_names = sorted(list(column_names))

    # Write column headers to the first row
    for col_idx, col_name in enumerate(column_names, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Starting from the second row, write record data
    row_idx = 2
    for record in second_level.findall("record"):
        # Collect all column values for the current record
        record_data = {}
        for column in record.findall("column"):
            if "name" in column.attrib:
                col_name = column.attrib["name"]
                # Handle empty text values
                col_value = column.text if column.text else ""
                record_data[col_name] = col_value

        # Write the record data to the corresponding cells
        for col_idx, col_name in enumerate(column_names, 1):
            if col_name in record_data:
                ws.cell(row=row_idx, column=col_idx, value=record_data[col_name])

        # Move to the next row for the next record
        row_idx += 1

# Save the Excel workbook to the output path
wb.save(excel_path)

# Notify user of successful completion
print(f"File Excel creato con successo: {excel_output}")
