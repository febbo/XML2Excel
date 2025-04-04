import os
import re
import xml.etree.ElementTree as ET
from openpyxl import Workbook


def repair_xml_entities(xml_path):
    """
    Repairs an XML file by replacing invalid characters and HTML entities with valid XML entities.

    The function works by:
    1. Reading the source XML file
    2. Finding content within <column> tags
    3. Replacing problematic characters and entities
    4. Saving the repaired content to a new file

    Args:
        xml_path (str): Path to the XML file that needs repair

    Returns:
        str: Path to the repaired XML file
    """
    base, ext = os.path.splitext(xml_path)
    output_path = f"{base}_repaired{ext}"

    with open(xml_path, "r", encoding="utf-8", errors="replace") as file:
        content = file.read()

    # Regular expression to match column tags and their content
    pattern = r"(<column[^>]*>)(.*?)(<\/column>)"

    def replace_entities(match):
        """
        Helper function to process each matched column tag and its content.
        Replaces HTML entities and special characters with valid XML entities.

        Args:
            match: A regex match object containing the column tag and its content

        Returns:
            str: The processed tag with properly escaped content
        """
        tag_open = match.group(1)
        content = match.group(2)
        tag_close = match.group(3)

        # Replace angle brackets with XML entities
        content = content.replace("<", "&lt;")
        content = content.replace(">", "&gt;")

        # Replace common HTML entities with plain text equivalents
        content = content.replace("&nbsp;", " ")
        content = content.replace("&ndash;", "-")
        content = content.replace("&mdash;", "-")
        content = content.replace("&copy;", "(c)")
        content = content.replace("&reg;", "(r)")
        content = content.replace("&trade;", "(tm)")
        content = content.replace("&lsquo;", "'")
        content = content.replace("&rsquo;", "'")
        content = content.replace("&ldquo;", '"')
        content = content.replace("&rdquo;", '"')
        content = content.replace("&bull;", "*")
        content = content.replace("&hellip;", "...")
        content = content.replace("&prime;", "'")
        content = content.replace("&Prime;", '"')
        content = content.replace("&frasl;", "/")
        content = content.replace("&euro;", "EUR")
        content = content.replace("&pound;", "GBP")
        content = content.replace("&yen;", "JPY")

        # Temporarily protect standard XML entities
        content = content.replace("&amp;", "&_amp_temp;")
        content = content.replace("&lt;", "&_lt_temp;")
        content = content.replace("&gt;", "&_gt_temp;")
        content = content.replace("&quot;", "&_quot_temp;")
        content = content.replace("&apos;", "&_apos_temp;")

        # Remove any remaining unrecognized entities
        content = re.sub(r"&[a-zA-Z0-9#]+;", "", content)

        # Restore the protected XML entities
        content = content.replace("&_amp_temp;", "&amp;")
        content = content.replace("&_lt_temp;", "&lt;")
        content = content.replace("&_gt_temp;", "&gt;")
        content = content.replace("&_quot_temp;", "&quot;")
        content = content.replace("&_apos_temp;", "&apos;")

        # Replace any remaining standalone ampersands with &amp;
        content = re.sub(r"&(?!(amp;|lt;|gt;|quot;|apos;))", "&amp;", content)

        return tag_open + content + tag_close

    repaired_content = re.sub(pattern, replace_entities, content, flags=re.DOTALL)

    with open(output_path, "w", encoding="utf-8") as file:
        file.write(repaired_content)

    print(f"Repaired XML file saved as: {output_path}")
    return output_path


def main():
    """
    Main function that orchestrates the XML repair and Excel conversion process.

    The function:
    1. Prompts the user for the XML filename
    2. Repairs the XML file to ensure it's well-formed
    3. Parses the repaired XML file
    4. Creates an Excel workbook with sheets for each second-level element in the XML
    5. Populates each sheet with data from the XML records
    """
    # Get the directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Prompt user for XML file name
    xml_file = input(
        "Enter the XML file name (must be in the same folder) with extension xml: "
    )
    xml_path = os.path.join(script_dir, xml_file)

    # Validate file existence
    if not os.path.isfile(xml_path):
        print(f"Error: The file {xml_file} does not exist in the folder {script_dir}")
        return

    try:
        print("Repairing the XML file...")
        repaired_xml_path = repair_xml_entities(xml_path)

        # Prepare Excel output filename
        excel_output = input(
            "Enter the output Excel file name (or press Enter to use the same name): "
        ).strip()
        if not excel_output:
            excel_output = (
                os.path.splitext(os.path.basename(repaired_xml_path))[0] + ".xlsx"
            )
        elif not excel_output.lower().endswith(".xlsx"):
            excel_output += ".xlsx"

        excel_path = os.path.join(script_dir, excel_output)

        print("Parsing the repaired XML file...")
        parser = ET.XMLParser(encoding="utf-8")
        tree = ET.parse(repaired_xml_path, parser)
        root = tree.getroot()

        # Create a new workbook and remove the default sheet
        wb = Workbook()
        wb.remove(wb.active)

        print("Creating Excel sheets...")

        # Process each second-level element in the XML structure
        for second_level in root:
            # Use element tag as sheet name, limiting to Excel's 31-character max
            sheet_name = second_level.tag[:31]

            # Handle duplicate sheet names by adding a counter
            counter = 1
            base_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name}_{counter}"[:31]
                counter += 1

            ws = wb.create_sheet(title=sheet_name)

            # Collect all unique column names from the records
            column_names = set()
            for record in second_level.findall("record"):
                for column in record.findall("column"):
                    if "name" in column.attrib:
                        column_names.add(column.attrib["name"])

            # Sort column names for consistent output
            column_names = sorted(list(column_names))

            # Write header row with column names
            for col_idx, col_name in enumerate(column_names, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Start data rows from row 2 (after header)
            row_idx = 2
            record_count = 0

            # Process each record and write to Excel
            for record in second_level.findall("record"):
                record_count += 1

                # Extract data from each column in the record
                record_data = {}
                for column in record.findall("column"):
                    if "name" in column.attrib:
                        col_name = column.attrib["name"]
                        col_value = column.text if column.text else ""
                        record_data[col_name] = col_value

                # Write record data to the appropriate cells
                for col_idx, col_name in enumerate(column_names, 1):
                    if col_name in record_data:
                        ws.cell(
                            row=row_idx, column=col_idx, value=record_data[col_name]
                        )

                row_idx += 1

            print(
                f"Sheet '{sheet_name}' created with {record_count} records and {len(column_names)} columns"
            )

        # Save the completed Excel workbook
        wb.save(excel_path)
        print(f"Excel file successfully created: {excel_output}")

    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    main()
