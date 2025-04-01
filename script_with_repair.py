import os
import re
import xml.etree.ElementTree as ET
from openpyxl import Workbook


def repair_xml_entities(xml_path):
    """
    Repair an XML file by replacing the <, > characters and common HTML entities with valid XML entities
    """
    base, ext = os.path.splitext(xml_path)
    output_path = f"{base}_repaired{ext}"

    with open(xml_path, "r", encoding="utf-8", errors="replace") as file:
        content = file.read()

    pattern = r"(<column[^>]*>)(.*?)(<\/column>)"

    def replace_entities(match):
        tag_open = match.group(1)
        content = match.group(2)
        tag_close = match.group(3)

        content = content.replace("<", "&lt;")
        content = content.replace(">", "&gt;")

        content = content.replace("&nbsp;", " ")

        content = content.replace("&ndash;", "-")
        content = content.replace("&mdash;", "-")
        content = content.replace("&copy;", "(c)")
        content = content.replace("&reg;", "(r)")
        content = content.replace("&trade;", "(tm)")
        content = content.replace("&lsquo;", "'")
        content = content.replace("&rsquo;", "'")
        content = content.replace("&ldquo;", '"')
        content = content.replace("&rdquo;", '"')
        content = content.replace("&bull;", "*")
        content = content.replace("&hellip;", "...")
        content = content.replace("&prime;", "'")
        content = content.replace("&Prime;", '"')
        content = content.replace("&frasl;", "/")
        content = content.replace("&euro;", "EUR")
        content = content.replace("&pound;", "GBP")
        content = content.replace("&yen;", "JPY")

        content = content.replace("&amp;", "&_amp_temp;")
        content = content.replace("&lt;", "&_lt_temp;")
        content = content.replace("&gt;", "&_gt_temp;")
        content = content.replace("&quot;", "&_quot_temp;")
        content = content.replace("&apos;", "&_apos_temp;")

        content = re.sub(r"&[a-zA-Z0-9#]+;", "", content)

        content = content.replace("&_amp_temp;", "&amp;")
        content = content.replace("&_lt_temp;", "&lt;")
        content = content.replace("&_gt_temp;", "&gt;")
        content = content.replace("&_quot_temp;", "&quot;")
        content = content.replace("&_apos_temp;", "&apos;")

        content = re.sub(r"&(?!(amp;|lt;|gt;|quot;|apos;))", "&amp;", content)

        return tag_open + content + tag_close

    repaired_content = re.sub(pattern, replace_entities, content, flags=re.DOTALL)

    with open(output_path, "w", encoding="utf-8") as file:
        file.write(repaired_content)

    print(f"File XML riparato salvato come: {output_path}")
    return output_path


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xml_file = input(
        "Inserisci il nome del file XML (deve essere nella stessa cartella): "
    )
    xml_path = os.path.join(script_dir, xml_file)

    if not os.path.isfile(xml_path):
        print(f"Errore: Il file {xml_file} non esiste nella cartella {script_dir}")
        return

    try:
        print("Riparazione del file XML in corso...")
        repaired_xml_path = repair_xml_entities(xml_path)

        excel_output = (
            os.path.splitext(os.path.basename(repaired_xml_path))[0] + ".xlsx"
        )
        excel_path = os.path.join(script_dir, excel_output)

        print("Parsing del file XML riparato...")
        parser = ET.XMLParser(encoding="utf-8")
        tree = ET.parse(repaired_xml_path, parser)
        root = tree.getroot()

        wb = Workbook()
        wb.remove(wb.active)

        print("Creazione dei fogli Excel...")

        for second_level in root:
            sheet_name = second_level.tag[:31]

            counter = 1
            base_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name}_{counter}"[:31]
                counter += 1

            ws = wb.create_sheet(title=sheet_name)

            column_names = set()
            for record in second_level.findall("record"):
                for column in record.findall("column"):
                    if "name" in column.attrib:
                        column_names.add(column.attrib["name"])

            column_names = sorted(list(column_names))

            for col_idx, col_name in enumerate(column_names, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            row_idx = 2
            record_count = 0
            for record in second_level.findall("record"):
                record_count += 1

                record_data = {}
                for column in record.findall("column"):
                    if "name" in column.attrib:
                        col_name = column.attrib["name"]
                        col_value = column.text if column.text else ""
                        record_data[col_name] = col_value

                for col_idx, col_name in enumerate(column_names, 1):
                    if col_name in record_data:
                        ws.cell(
                            row=row_idx, column=col_idx, value=record_data[col_name]
                        )

                row_idx += 1

            print(
                f"Foglio '{sheet_name}' creato con {record_count} record e {len(column_names)} colonne"
            )

        wb.save(excel_path)
        print(f"File Excel creato con successo: {excel_output}")

    except Exception as e:
        print(f"Si è verificato un errore: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    main()
